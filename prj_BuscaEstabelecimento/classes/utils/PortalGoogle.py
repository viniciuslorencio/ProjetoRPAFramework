from prj_BuscaEstabelecimento.classes.framework.InitAllSettings import InitAllSettings as InitAllSetting
from prj_BuscaEstabelecimento.classes.framework.GetTransaction import GetTransaction as GetTransaction
from prj_BuscaEstabelecimento.classes.utils.Maestro import Maestro as Maestro, LogLevel, ErrorType
from prj_BuscaEstabelecimento.classes.utils.Exceptions import BusinessRuleException
from prj_BuscaEstabelecimento.classes.utils.Utils import Auxiliar
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from pathlib import Path
import pandas as pd
import time


ROOT_DIR = Path(__file__).parent.parent.parent
class Maps:
    """
    Classe responsável pelo download de documentos no Ariba.

    Parâmetros:
    
    Retorna:
    """

    # Variáveis de classe
    var_dictConfig = InitAllSetting.var_dictConfig
    var_botWebbot = InitAllSetting.var_botWebbot

    @classmethod
    def abertura_portal(cls):
        """
        Realiza abertura de portal.

        Parâmetros:
        - driver: WebDriver Selenium instanciado.

        Retorna:
        None
        """
        try:
            Maestro.write_log(f"Iniciou abertura do Google Maps.")
            
            # Atualiza a variável cls.var_botWebbot com uma nova instância
            cls.var_botWebbot = InitAllSetting.var_botWebbot 

            # Abre a URL desejada
            var_AberturaSite = cls.var_botWebbot.get(cls.var_dictConfig["URL_Google"])
            cls.var_botWebbot.maximize_window()
            Maestro.write_log(f"Realizou a abertura do site:{cls.var_dictConfig["URL_Google"]}")
            time.sleep(10)
            
        except Exception as e:
            Maestro.write_log(f"Erro ao realizar abertura: {e}")
            raise Exception
                    
    @classmethod
    def pesquisar_item(cls):
        """
        Realiza o coleta de dados.

        Parâmetros:
        - driver: WebDriver Selenium instanciado.

        Retorna:
        None
        """
        try:
            Maestro.write_log(f"Iniciou a pesquisa por viagens.")

            var_strTipoBusca = (GetTransaction.var_dictQueueItem['Tipo'] +" "+GetTransaction.var_dictQueueItem['Cidade'])
            var_wePreencherTipo = cls.var_botWebbot.find_element(By.XPATH, '//form/input[@class="fontBodyMedium searchboxinput xiQnY "]')
            var_wePreencherTipo.send_keys(Keys.CONTROL + 'a')
            var_wePreencherTipo.send_keys(var_strTipoBusca)
            Maestro.write_log(f"Preencheu pesquisa com: {var_strTipoBusca}.")
            
            time.sleep(2.5)
            
            var_weClickPesquisar = cls.var_botWebbot.find_element(By.XPATH, '//button[@id="searchbox-searchbutton"]').click()
            Maestro.write_log(f"Clicou em pesquisar.")

            # Agurdar retorno da pesquisa
            var_TempoMaximo = 300  
            start_time = time.time()
            var_auxiliar = False

            while True:
                try:
                    # Tenta localizar o elemento
                    if cls.var_botWebbot.find_elements(By.XPATH, f'//div[@aria-label="{var_strTipoBusca}"]') or cls.var_botWebbot.find_elements(By.XPATH, f'//div[@aria-label="Resultados para {var_strTipoBusca}"]') :
                        time.sleep(1.5)
                        Maestro.write_log(f"A pesquisa {var_strTipoBusca} retornou resultado.")
                        var_auxiliar = True
                        return var_auxiliar
                    
                except Exception:
                    # Se não encontrar o elemento, aguarda 1.5 segundos e tenta novamente
                    time.sleep(2)
                    elapsed_time = time.time() - start_time
                    if elapsed_time > var_TempoMaximo:
                        raise Exception("Tempo máximo de espera excedido para aguardar resultado.")

        except Exception as e:
            Maestro.write_log(f"Erro ao realizar pesquisa: {e}")
            # raise BusinessRuleException(f"Erro no processo PortalUnico: {str(e)}")
            
            raise Exception(f"Erro no processo PortalUnico: {str(e)}")

    @classmethod
    def percorre_lista(cls):
        """
        Realiza o download de documentos no Ariba.

        Parâmetros:
        - driver: WebDriver Selenium instanciado.

        Retorna:
        None
        """
        try:
            Maestro.write_log(f"Iniciou captura de valores.")
            qty_raw = GetTransaction.var_dictQueueItem.get('Quantidade', 0)
            quantidade = int(qty_raw)
            var_listResultados = []


            for i in range(1, quantidade + 1):  # XPath usa índice começando em 1
                var_weListaDados = cls.var_botWebbot.find_element(By.XPATH, f'(//div/a[@class="hfpxzc"])[{i}]')
                
                for _ in range(3):
                    var_weListaDados.send_keys(Keys.ARROW_DOWN)
                    time.sleep(0.08)
                
                var_weListaDados.click()
                var_weListaDados.send_keys(Keys.ARROW_DOWN)

                time.sleep(3)

                # Agurdar retorno da pesquisa
                var_TempoMaximo = 300  
                start_time = time.time()
                while True:
                    try:
                        # Tenta localizar o elemento
                        if cls.var_botWebbot.find_elements(By.XPATH, '//h1[@class="DUwDvf lfPIob"]'):
                            Maestro.write_log(f"Conteudo carregado.")
                            break
                        else:
                            time.sleep(3)
                            Maestro.write_log(f"Aguardando carregamento de conteudo.")
                            
                    except Exception:
                        # Se não encontrar o elemento, aguarda 1.5 segundos e tenta novamente
                        time.sleep(2)
                        elapsed_time = time.time() - start_time
                        if elapsed_time > var_TempoMaximo:
                            raise Exception("Tempo máximo de espera excedido para aguardar resultado.")

                var_strNomeEmpresa = cls.var_botWebbot.find_element(By.XPATH, '//h1[@class="DUwDvf lfPIob"]').text
                Maestro.write_log(f"Capturou o nome do estabelicimento: {var_strNomeEmpresa}")
                
                var_strClassificacao = cls.var_botWebbot.find_element(By.XPATH, '//div[@class="fontBodyMedium dmRWX"]').text
                linhas = [l.strip() for l in var_strClassificacao.splitlines() if l.strip()]
                var_strQtnEstrelas = linhas[0] if len(linhas) >= 1 else ''
                Maestro.write_log(f"A quantidade de estrelas são: {var_strQtnEstrelas}")

                var_strQtnAvaliacoes = linhas[1].strip('()') if len(linhas) >= 2 else ''
                Maestro.write_log(f"A quantidade de usuarios que avaliaram são: {var_strQtnAvaliacoes}")

                var_strEstrutura= cls.var_botWebbot.find_element(By.XPATH, '(//div[@class="m6QErb XiKgde "])[1]').text
                
                var_strEndereco,var_strTelefone,var_strLink = Auxiliar.extrai_dados(var_strEstrutura)
                Maestro.write_log(f"O endereço obtido: {var_strEndereco}")
                Maestro.write_log(f"Numero de telefone obtido: {var_strTelefone}")
                Maestro.write_log(f"Link do site obtido: {var_strLink}")

                var_listResultados.append({
                    'Nome': var_strNomeEmpresa,
                    'Estrelas': var_strQtnEstrelas,
                    'Avaliacoes': var_strQtnAvaliacoes,
                    'Endereco': var_strEndereco,
                    'Telefone': var_strTelefone,
                    'Site': var_strLink
                })

            df = pd.DataFrame(var_listResultados)

            excel_file = InitAllSetting.var_strCaminhoResultadosXLSX

            # obtém o nome da sheet — ajuste a chave se você passar outro nome
            sheet_name = GetTransaction.var_dictQueueItem.get('SheetName') or str(GetTransaction.var_dictQueueItem.get('Tipo', 'Resultados'))

            # abre workbook e garante que não vamos sobrescrever uma sheet existente
            wb = load_workbook(excel_file)
            Maestro.write_log(f"Realizou o carregamento da planilha: {excel_file}")

            target_sheet = sheet_name
            if target_sheet in wb.sheetnames:
                base = sheet_name
                i = 1
                while f"{base}_{i}" in wb.sheetnames:
                    i += 1
                target_sheet = f"{base}_{i}"
            # escreve em modo append criando uma nova sheet
            with pd.ExcelWriter(excel_file, engine="openpyxl", mode="a") as writer:
                df.to_excel(writer, sheet_name=target_sheet, index=False)
                Maestro.write_log(f"Dados salvos com sucesso.")

            # 3) reabre e agora sim apaga Plan1, se houver outra aba
            wb = load_workbook(excel_file)
            if "Plan1" in wb.sheetnames and len(wb.sheetnames) > 1:
                del wb["Plan1"]
                wb.save(excel_file)    

            Maestro.write_log(f"Planilha gravada: {excel_file} (sheet: {target_sheet})")

            var_Status = "SUCESSO"
            var_Observacao = ""
            Maestro.write_log(f"Finalizou relatorio de estabelecimentos.")
            return var_Status, var_Observacao
        
        except Exception as e:
            Maestro.write_log(f"Erro ao realizar pesquisa: {e}")
            raise BusinessRuleException(f"Erro no processo PortalUnico: {str(e)}")

    