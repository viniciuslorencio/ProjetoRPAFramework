# Imports dos modulos  
# Carrega o InitAllSettingsSettings Precisa ser o primeiro a ser carregado
from prj_BuscaEstabelecimento.classes.framework.InitAllSettings import InitAllSettings as InitAllSettings
from prj_BuscaEstabelecimento.classes.utils.Maestro import Maestro as Maestro, LogLevel, ErrorType

# Imports dos pacotes externos
from openpyxl import load_workbook
from datetime import datetime
import pandas as pd
import shutil, os
import platform
import sqlite3
import re



class Relatorios:
    """
    Classe responsável por manipular relatórios, incluindo linhas e preenchendo a partir de templates
    
    Parâmetros:

    Retorna:

    """
    var_dictConfig = InitAllSettings.var_dictConfig
    var_strGuidExecucao = InitAllSettings.var_strGuidExecucao
    var_strCaminhoBDAnalitSint = InitAllSettings.var_strSQLiteCaminhoBDAnalitSint
    
    @classmethod
    def preencher_analitico(cls,var_IdsProcessados) -> str:
        """
        Realiza o preenchimento do excel do relatorio analitico.
        
        Parâmetros:
            cls.var_strPathRelatorioAnalitico (str): Caminho da planilha analítica a ser preenchida.
            InitAllSettings.var_strCaminhoBancoXlsx (str): Caminho da planilha banco_dados.xlsx.
            var_IdsProcessados (list): Lista de IDs que foram processados.
        
        Retorna:
            - str: Caminho do arquivo preenchido com os dados do relatorio analitico.
        
        """
        cls.var_strPathRelatorioAnalitico:str = os.path.join(cls.var_dictConfig["CaminhoPastaRelatorios"] , "Relatorio_Analitico_" + re.sub(string=str(cls.var_dictConfig["NomeProcesso"]),pattern=r'[\@\$%&\\\/\:\*\?\"\'<>\|~`#\^\+=\{\}\[\];\!]',repl='') + "___datetime__" + ".xlsx")
        cls.var_strPathRelatorioAnalitico = cls.var_strPathRelatorioAnalitico.replace("__datetime__",datetime.now().strftime("%d%m%Y"))

        #Copiando apenas se não existir
        if(not os.path.exists(cls.var_strPathRelatorioAnalitico)):
            shutil.copy(src=InitAllSettings.var_strCaminhoTemplateExcelAnalitico, dst=cls.var_strPathRelatorioAnalitico)
        
            #Colocando nome do processo no relatório
            var_wbkAnalitico = load_workbook(cls.var_strPathRelatorioAnalitico)
            var_wshtAnalitico = var_wbkAnalitico.active
            var_wshtAnalitico["D4"] = cls.var_dictConfig["NomeProcesso"]

            var_wbkAnalitico.save(cls.var_strPathRelatorioAnalitico)
            var_wbkAnalitico.close()


        # Caminho do CSV
        caminho_csv = InitAllSettings.var_strCaminhoBancoCSV

        # Lê o CSV assumindo que há cabeçalho e os IDs estão na primeira coluna
        df_banco = pd.read_csv(caminho_csv)

        # Nome da primeira coluna (onde estão os IDs)
        nome_coluna_id = df_banco.columns[0]

        # Filtra as linhas cujos IDs estão em var_IdsProcessados (retorna a linha inteira)
        df_filtrado = df_banco[df_banco[nome_coluna_id].isin(var_IdsProcessados)]
            
        # Abrir a planilha analítica
        wb_analitica = load_workbook(cls.var_strPathRelatorioAnalitico)
        aba_analitica = wb_analitica["2. Analítico"]
        
        # Encontrar a primeira linha vazia na coluna A
        linha_vazia = 5  # Assume que os dados começam na linha 5
        while aba_analitica[f"A{linha_vazia}"].value is not None:
            linha_vazia += 1
            
        # Inserir os dados filtrados na planilha analítica
        for _, row in df_filtrado.iterrows():
            aba_analitica[f"A{linha_vazia}"] = row["Data iniciado"]
            aba_analitica[f"B{linha_vazia}"] = row["Data finalizado"]
            aba_analitica[f"C{linha_vazia}"] = row["Tipo"]
            aba_analitica[f"D{linha_vazia}"] = row["Cidade"]
            aba_analitica[f"E{linha_vazia}"] = row["Quantidade"]
            aba_analitica[f"F{linha_vazia}"] = row["Caminho da Pasta"]
            aba_analitica[f"G{linha_vazia}"] = row["Status"]
            aba_analitica[f"H{linha_vazia}"] = row["Obs"]
            linha_vazia += 1
            
        # Salvar as alterações na planilha analítica
        wb_analitica.save(cls.var_strPathRelatorioAnalitico)
        wb_analitica.close()                 
        return cls.var_strPathRelatorioAnalitico 
    
    @classmethod    
    def preencher_sintetico(cls,var_IdsProcessados) -> str:
        """
        Realiza o preenchimento do excel do relatorio sintetico.
        
        Parâmetros:
            var_IdsProcessados (list): Lista de IDs processados durante a execução.
        
        Retorna:
            - str: Caminho do arquivo preenchido com os dados do relatorio sintetico.
        
        """
        cls.var_strPathRelatorioSintetico:str = os.path.join(cls.var_dictConfig["CaminhoPastaRelatorios"] , "Relatorio_Sintetico_" + re.sub(string=str(cls.var_dictConfig["NomeProcesso"]),pattern=r'[\@\$%&\\\/\:\*\?\"\'<>\|~`#\^\+=\{\}\[\];\!]',repl='') + "___datetime__" + ".xlsx")
        cls.var_strPathRelatorioSintetico = cls.var_strPathRelatorioSintetico.replace("__datetime__",datetime.now().strftime("%m%Y"))


        # Captura modelo de arquivo para preencher o relatorio sintetico 
        if(not os.path.exists(cls.var_strPathRelatorioSintetico)):
            shutil.copy(src=InitAllSettings.var_strCaminhoTemplateExcelSintetico, dst=cls.var_strPathRelatorioSintetico)
        
            #Colocando nome do processo no relatório
            var_wbkSintetico = load_workbook(cls.var_strPathRelatorioSintetico)
            var_wshtSintetico = var_wbkSintetico.active
            var_wshtSintetico["C4"] = cls.var_dictConfig["NomeProcesso"]

            var_wbkSintetico.save(cls.var_strPathRelatorioSintetico)
            var_wbkSintetico.close()

        # Dados fornecidos
        nome_processo = cls.var_dictConfig["NomeProcesso"]
        data_hora_inicio = InitAllSettings.var_strDatahoraInicioExec
        data_hora_fim = InitAllSettings.var_strDatahoraFimExec
        quantidade_itens_processados = len(var_IdsProcessados)
        maquina_executada = platform.node()


        # Carregar o relatório sintético
        wb_sintetico = load_workbook(cls.var_strPathRelatorioSintetico)
        aba_sintetico = wb_sintetico.active

        # Encontrar a primeira linha vazia na coluna A (assumindo que a primeira coluna tem dados únicos)
        linha_vazia = 6  # Supondo que a primeira linha seja o cabeçalho
        while aba_sintetico[f"A{linha_vazia}"].value is not None:
            linha_vazia += 1

        # Calcular o total de horas
        dt_inicio = datetime.strptime(data_hora_inicio, "%d/%m/%Y %H:%M:%S")
        dt_fim = datetime.strptime(data_hora_fim, "%d/%m/%Y %H:%M:%S")
        total_horas = dt_fim - dt_inicio

        # Preencher as colunas do relatório
        aba_sintetico[f"A{linha_vazia}"] = nome_processo
        aba_sintetico[f"B{linha_vazia}"] = data_hora_inicio
        aba_sintetico[f"C{linha_vazia}"] = data_hora_fim
        aba_sintetico[f"D{linha_vazia}"] = str(total_horas)
        aba_sintetico[f"E{linha_vazia}"] = quantidade_itens_processados
        aba_sintetico[f"F{linha_vazia}"] = maquina_executada

        # Salvar as alterações no relatório
        wb_sintetico.save(cls.var_strPathRelatorioSintetico)
        wb_sintetico.close()

        return cls.var_strPathRelatorioSintetico


























































































































        # --------------------------------------------------------------------------------------------------------------------------------------------
        #Script para realizar a captura dos dados para relatorio analitico
        with open(InitAllSettings.var_strCaminhoScriptSelectDadosSintetico) as script_file:
            var_strScriptCapturaDadosSintetico = script_file.read()
        
        # Realizando a conexão do banco
        var_csrCursor = sqlite3.connect(cls.var_strCaminhoBDAnalitSint)
        
        # Executando o comando select
        var_listDadosSintetico = var_csrCursor.execute(var_strScriptCapturaDadosSintetico,[cls.var_strGuidExecucao]).fetchall()

        # Nome aba arquivo excel que será realizada a leitura
        var_strSheetName = '1. Sintético'

        # Carregar o arquivo Excel
        var_wbExcelSintetico = load_workbook(cls.var_strPathRelatorioSintetico)

        # Selecionar a aba Sintetico
        var_wsAbaSintetico= var_wbExcelSintetico[var_strSheetName]

        # Variaveis auxiliares para encontrar linha vazia no excel                
        var_intIndexNewline:int = None
        var_intIndexAux = 5

        # Encontrando linha vazia
        while(var_intIndexNewline is None):
            if(var_wsAbaSintetico["A" + var_intIndexAux.__str__()].value is None):
                var_intIndexNewline = var_intIndexAux
            else:
                var_intIndexAux += 1

        # Percorre linhas e colunas do excel para preencher com os dados sinteticos
        for r_idx, row in enumerate(var_listDadosSintetico, 1):
            for c_idx, value in enumerate(row, 1):
                var_wsAbaSintetico.cell(row=r_idx+var_intIndexNewline-1, column=c_idx, value=value)

        # Salva as modificações
        var_wbExcelSintetico.save(cls.var_strPathRelatorioSintetico)

        # Fecha a conexão com o SQLite
        var_csrCursor.close()
        
        return cls.var_strPathRelatorioSintetico     
    
    

    

